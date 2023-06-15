from ast import While
from dis import dis
from itertools import count
from math import prod
from re import L, X
from tkinter import Y
from xml.dom import INVALID_ACCESS_ERR
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import styles
import numpy as np
from datetime import datetime, date
from glob import glob
import pandas as pd
from tkinter import filedialog
from fileinput import filename
import os
from cProfile import run

loaded_files = []

#load in workbook and active worksheet from .xlsx
#Allow user to select each report
def load_cw_report():
    ConnectWise = filedialog.askopenfilename(initialdir='./', title="Select Report", filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
    loaded_files.append(ConnectWise)
    print(ConnectWise)   

def load_qb_report():
    QuickBooks = filedialog.askopenfilename(initialdir='./', title="Select Report", filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
    loaded_files.append(QuickBooks)
    print(QuickBooks)

def run_report():    
    wb_cw = load_workbook(loaded_files[0])
    wb_qb = load_workbook(loaded_files[1])

    #wb_cw = load_workbook('ConnectWise.xlsx')
    #wb_qb = load_workbook('QuickBooks.xlsx')

    ws_cw = wb_cw.active
    ws_qb = wb_qb.active

    #Create workbook and sheets for data results
    wb_results = Workbook('Results.xlsx')
    ws_compared = wb_results.create_sheet(title = 'Accurate Count')
    ws_cw_only = wb_results.create_sheet(title = 'ConnectWise Only')
    ws_qb_only = wb_results.create_sheet(title = 'QuickBooks Only')

    wb_results.save('Results.xlsx')

    wb_r = load_workbook('Results.xlsx')
    ws_compare = wb_r['Accurate Count']
    ws_cw_ = wb_r['ConnectWise Only']
    ws_qb_ = wb_r['QuickBooks Only']
    #ws2 = wb.create_sheet('Sheet2')
    #ws3 = wb.create_sheet('Count issues')
    #print(wb.sheetnames)

    #write to specific cell
    #ws1['A1'].value = "ConnectWise"
    #ws1['A9'].value = "test"
    #ws1['E9'].value = "test"
    #ws1['C9'].value = 3
    #ws1['G9'].value = 3
    #ws1['A10'].value = "test"
    #ws1['E10'].value = "test"
    #ws1['C10'].value = 5
    #ws1['G10'].value = 4

    #declare arrays to store data
    product_id_a = []
    product_id_e = []
    cw_count = []
    qb_count = []

    #arrays for combined data from above arrays
    cw = []
    qb = []

    #pull all data from column of data
    for cell_a in ws_cw['B']: # test: 'A' | final: 'B'
        #stop looping if there is no more data in columns
        if cell_a.value == None:
            no_val = 0
            while no_val > 10:
                break
            no_val += 1
        #add data to declared array
        product_id_a.append(cell_a.value)
        
    #pull all data from column of data
    for cell_c in ws_cw['L']: # test: 'C' | final: 'J'
        #stop looping if there is no more data in columns
        if cell_c.value == None:
            #some values in this array may have no data. Loop 'x' number of times before deciding end of data in column
            no_val = 0
            while no_val > 5:
                break
            no_val += 1
        #add data to declared array
        cw_count.append(cell_c.value)

    #pull all data from column of data
    for cell_e in ws_qb['A']: # test: 'E' | final: 'L'
        #stop looping if there is no more data in columns
        if cell_e.value == None:
            break
        #add data to declared array
        product_id_e.append(cell_e.value)

    #pull all data from column of data
    for cell_g in ws_qb['K']: # test: 'G' | final: 'N'
        #stop looping if there is no more data in columns
        if cell_g.value == None:
            #Some values in this array may have no data. Loop 'x' number of times before deciding end of data in column
            no_val = 0
            while no_val > 5:
                break
            no_val += 1
        #add data to declared array
        qb_count.append(cell_g.value)

    #compare column1[x] (product_id_a[x]) with column2[(a-z)] (product_id_e[x])
    cw_only = []
    in_both = []
    for compare_id_cw in product_id_a:
        #determine similarities between ConnectWise and Quickbooks
        if compare_id_cw in product_id_e:
            if compare_id_cw in product_id_e:
                in_both.append(compare_id_cw)
        #determine differences between ConnectWise and Quickbooks
        if compare_id_cw not in product_id_e:
            if compare_id_cw not in cw_only:
                cw_only.append(compare_id_cw)

    ws_cw_['D1'].value = 'Full ConnectWise ID list'
    #combine data from ID column and Count column (ConnectWise)
    for pos_cw in range(len(product_id_a)):
        x = product_id_a[pos_cw], str(cw_count[pos_cw]), pos_cw
        cw.append(x)
        ws_cw_.cell(row=3 + pos_cw, column=4).value = str(product_id_a[pos_cw])
        ws_cw_.cell(row=3 + pos_cw, column=5).value = str(cw_count[pos_cw])

        #Writes data only in ConnectWise to columns A and B in CW Only sheet
        while product_id_a[pos_cw] not in product_id_e:
            ws_cw_.cell(row = 3 + pos_cw, column=1).value = str(product_id_a[pos_cw])
            ws_cw_.cell(row = 3 + pos_cw, column=2).value = str(cw_count[pos_cw])
            break
    ws_cw_['A1'].value = 'ConnectWise Only'

    ws_qb_['D1'].value = 'QuickBooks ID list'
    #combine data from ID column and Count column (QuickBooks)
    for pos_qb in range(len(product_id_e)):
        y = product_id_e[pos_qb], str(qb_count[pos_qb]), pos_qb
        qb.append(y)
        ws_qb_.cell(row=3 + pos_qb, column=4).value = str(product_id_e[pos_qb])
        ws_qb_.cell(row=3 + pos_qb, column=5).value = str(qb_count[pos_qb])

        #Writes data only in QuickBooks to columns A and B in QB Only sheet
        while product_id_e[pos_qb] not in product_id_a:
            ws_qb_.cell(row=3 + pos_qb, column=1).value = str(product_id_e[pos_qb])
            ws_qb_.cell(row=3 + pos_qb, column=2).value = str(qb_count[pos_qb])
            break
    ws_qb_['A1'].value = 'QuickBooks Only'

    #compare column2[x] (product_id_e[x]) with column1[(a-z)]
    qb_only = []
    for compare_id_qb in product_id_e:
        if compare_id_qb not in product_id_a:
            if compare_id_qb not in qb_only:
                qb_only.append(compare_id_qb)

    diff_val_id = []
    diff_val_cw = []
    diff_val_qb = []
    diff_val_combine = []
    same_val_id = []
    same_val_count_cw = []
    same_val_count_qb = []

    for cw_1 in cw:
        for qb_1 in qb:
            if str(cw_1[0]) in str(qb_1[0]):
                if cw_1[1] not in qb_1[1]:
                    diff_val_id.append(str(cw_1[0]))
                    diff_val_cw.append(str(cw_1[1]))
                    diff_val_qb.append(str(qb_1[1]))

                    diff_val_combine.append(cw_1[0])
                    diff_val_combine.append(cw_1[1])
                    diff_val_combine.append(qb_1[1])
                if cw_1[1] in qb_1[1]:
                    same_val_id.append(str(cw_1[0]))
                    same_val_count_cw.append(str(cw_1[1]))
                    same_val_count_qb.append(str(qb_1[1]))

    ws_compare['A1'].value = "Incorrect Count"
    ws_compare['A2'].value = "Product ID"
    ws_compare['B2'].value = "ConnectWise"
    ws_compare['C2'].value = "QuickBooks"

    ws_compare['E1'].value = "Correct Count"
    ws_compare['E2'].value = "Product ID"
    ws_compare['F2'].value = "ConnectWise"
    ws_compare['G2'].value = "QuickBooks"

    for diff_id in range(len(diff_val_id)):
        ws_compare.cell(row=3 + diff_id, column=1).value = diff_val_id[diff_id]

    for diff_cw in range(len(diff_val_cw)):
        ws_compare.cell(row=3 + diff_cw, column=2).value = str(diff_val_cw[diff_cw])

    for diff_qb in range(len(diff_val_qb)):
        ws_compare.cell(row=3 + diff_qb, column=3).value = str(diff_val_qb[diff_qb])

    for same_id in range(len(same_val_id)):
        ws_compare.cell(row=3 + same_id, column=5).value = str(same_val_id[same_id])

    for same_cw in range(len(same_val_id)):
        ws_compare.cell(row=3 + same_cw, column=6).value = str(same_val_count_cw[same_cw])

    for same_qb in range(len(same_val_id)):
        ws_compare.cell(row=3 + same_qb, column=7).value = str(same_val_count_qb[same_qb])
    
    #save file
    wb_r.save("Inventory Results " + datetime.now().strftime("%m.%d.%Y") + ".xlsx")
    print("\nYour file is saved as:\n" + "Inventory Results " + datetime.now().strftime("%m.%d.%Y") + ".xlsx\nFinished\n")

def saved_report():
    arr = []
    arr.append("Inventory Results " + datetime.now().strftime("%m.%d.%Y") + ".xlsx")
    strarr = str(arr).replace('[','').replace(']','').replace('\'','')
    os.startfile(strarr)